# --// IMPORTS \\-- #
# -- Libraries -- #
import cv2 as cv
import pandas as pd
from tabulate import tabulate
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# -- TypeCasting -- #
from numpy import ndarray

# -- Classes -- #
from classes.Element import Element
from classes.Page import Page
from classes.Contour import Contour
from classes.Table import Table


# --// PROGRAMM \\-- #
def printAndSaveRectangleElements(page: Page, elements: list[Element], path: str, filename: str) -> None:
    img_to_draw: ndarray = page.image.copy()
    for e in elements:
        cv.rectangle(img_to_draw, (e.x1, e.y1), (e.x2, e.y2), (0, 255, 0), 2)

    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)


def printAndSaveRectangle(img: ndarray, rectangles: list[tuple[int, int, int, int]], color: tuple, thickness: int, path: str, filename: str) -> None:
    img_to_draw = img.copy()
    for r in rectangles:
        cv.rectangle(img_to_draw, (r[0], r[1]), (r[2], r[3]), color, thickness)
    cv.imwrite(f'{path}/{filename}.jpg', img_to_draw)


def printAndSaveSpecialBox(page: Page, image: ndarray, elements: list[Element], path: str, filename: str) -> None:
    img_to_draw: ndarray = image.copy()
    for e in elements:
        cv.rectangle(img_to_draw, (e.x1, e.y1), (e.x2, e.y2), (0, 255, 0), 2)
        cv.line(img_to_draw, (e.x1, e.y1), (e.x2, e.y2), (255, 0, 0), 2)
        cv.line(img_to_draw, (e.x1, e.y2), (e.x2, e.y1), (255, 0, 0), 2)
        cv.line(img_to_draw, (e.h_mid[0], e.h_mid[1]), (e.h_mid[2], e.h_mid[3]), (0, 0, 255), 2)
        cv.line(img_to_draw, (e.v_mid[0], e.v_mid[1]), (e.v_mid[2], e.v_mid[3]), (0, 0, 255), 2)

    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)


def printAndSaveFormattedSpecialBox(page: Page, image: ndarray, elements: list[Element], path: str, filename: str) -> None:
    img_to_draw: ndarray = image.copy()
    for e in elements:
        cv.rectangle(img_to_draw, (e.fx1, e.fy1), (e.fx2, e.fy2), (0, 255, 0), 1)
        cv.line(img_to_draw, (e.fx1, e.fy1), (e.fx2, e.fy2), (255, 0, 0), 1)
        cv.line(img_to_draw, (e.fx1, e.fy2), (e.fx2, e.fy1), (255, 0, 0), 1)
        cv.line(img_to_draw, (e.fh_mid[0], e.fh_mid[1]), (e.fh_mid[2], e.fh_mid[3]), (0, 0, 255), 1)
        cv.line(img_to_draw, (e.fv_mid[0], e.fv_mid[1]), (e.fv_mid[2], e.fv_mid[3]), (0, 0, 255), 1)

    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)


def printAndSaveContours(page: Page, contours: tuple, path: str, filename: str):
    img_to_draw: ndarray = page.image.copy()
    cv.drawContours(img_to_draw, contours, -1, (0, 255, 0), 1)
    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)


def printAndSaveContourObjects(page: Page, contours: list[Contour], path: str, filename: str):
    img_to_draw: ndarray = page.image.copy()
    for contour in contours:
        cv.drawContours(img_to_draw, [contour.coords], -1, (0, 255, 0), 1)
    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)


def printAndSaveLines(page: Page, table: Table, lines: list[tuple], color: tuple, thickness: int, path: str, filename: str):
    img_to_draw: ndarray = table.img_border.copy()
    for line in lines:
        cv.line(img_to_draw, (line[0], line[1]), (line[2], line[3]), color, thickness)
    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)

def printAndSaveLinesList(page: Page, lines: list[tuple], color: tuple, thickness: int, path: str, filename: str):
    img_to_draw: ndarray = page.image.copy()
    for line in lines:
        cv.line(img_to_draw, (line[0], line[1]), (line[2], line[3]), color, thickness)
    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', img_to_draw)


def saveImage(page: Page, image: ndarray, path: str, filename: str) -> None:
    cv.imwrite(f'{path}/{page.nr}_{filename}.jpg', image)


def consoleOutputTable(table):
    # Create an empty DataFrame with the appropriate headers
    rows = set()
    cols = set()

    # Collect all unique row and column indices
    for e in table.elements:
        rows.update(e.row)
        cols.update(e.column)

    df = pd.DataFrame(index=sorted(rows), columns=sorted(cols))

    # Initialize the cells as empty lists
    for row in df.index:
        for col in df.columns:
            df.at[row, col] = []

    # Fill the DataFrame with the elements
    for e in table.elements:
        for r in e.row:
            for c in e.column:
                df.at[r, c].append(e.text)

    # Optional: Convert the lists to strings to make the table more readable
    for row in df.index:
        for col in df.columns:
            df.at[row, col] = ' '.join(df.at[row, col])

    # Output the table with tabulate
    print(tabulate(df, headers='keys', tablefmt='grid'))


def excelOutputTable(elements: list[Element], path: str, filename: str) -> None:
    def setCellValue(rows: list[int], cols: list[int], value: str) -> None:
        # Check if the cell already has a value and prepend the new value
        current_value = ws.cell(row=rows[0], column=cols[0]).value
        if current_value:
            value = str(current_value) + " " + str(value)

        # Set the value only in the first cell of the range
        ws.cell(row=rows[0], column=cols[0], value=value)


    def mergeCells(rows: list[int], cols: list[int]) -> None:
        # Check if the cell needs to be merged
        if len(rows) > 1 or len(cols) > 1:
            start_row = min(rows)
            end_row = max(rows)
            start_col = min(cols)
            end_col = max(cols)
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)

            # Center the text vertically if the cells are merged
            if len(rows) > 1:
                ws.cell(row=start_row, column=start_col).alignment = Alignment(vertical='center')

            # Center the text vertically if the cells are merged in a column
            if len(cols) > 1:
                ws.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')

    def addCellBorder(min_row: int, max_row: int, min_col: int, max_col: int) -> None:
        # Define the border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add the border to all cells in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border


    # Create a new Workbook and activate the first worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Initialize variables for the table range
    min_row, max_row = float('inf'), 0
    min_col, max_col = float('inf'), 0

    # Iterate over the elements and insert them into the worksheet
    for cell in elements:
        rows = [r + 1 for r in cell.row]  # Increment row values by 1
        cols = [c + 1 for c in cell.column]  # Increment column values by 1
        value = cell.text

        # Update the table range
        min_row = min(min_row, min(rows))
        max_row = max(max_row, max(rows))
        min_col = min(min_col, min(cols))
        max_col = max(max_col, max(cols))

        setCellValue(rows, cols, value)
        mergeCells(rows, cols)
        addCellBorder(min_row, max_row, min_col, max_col)

    # Save the Excel file
    wb.save(f'{path}/{filename}.xlsx')